import openpyxl
import re
import os
import shutil
from pathlib import Path

INPUT_DIR = Path("result")
OUTPUT_DIR = Path("fixed")

# Mapping: E row -> correct C row reference in 'CW 2 SHEET'
# E6->C2, E7->C3, E8->C4, E9->C5, E10->C6, E11->C7, E12->C8, E13->C9, E14->C10, E15->C11, E16->C12, E17->C13
# (E11 was skipping C7, and E17 was a duplicate C11)
CORRECT_REFS = {
    11: 7,
    12: 8,
    13: 9,
    14: 10,
    15: 11,
    16: 12,
}
CLEAR_ROW = 17  # E17 should be cleared (duplicate/wrong)

RESULT_SHEET = "CW 2 RESULT"
SOURCE_SHEET = "CW 2 SHEET"


def fix_file(src_path: Path, dst_path: Path):
    wb = openpyxl.load_workbook(src_path)

    if RESULT_SHEET not in wb.sheetnames:
        print(f"  SKIP: sheet '{RESULT_SHEET}' not found in {src_path.name}")
        shutil.copy2(src_path, dst_path)
        return

    sheet = wb[RESULT_SHEET]

    for e_row, c_row in CORRECT_REFS.items():
        cell = sheet.cell(row=e_row, column=5)
        current = cell.value
        if current and SOURCE_SHEET in str(current):
            new_val = f"='{SOURCE_SHEET}'!C{c_row}"
            if current != new_val:
                cell.value = new_val

    # Clear the duplicate E17
    clear_cell = sheet.cell(row=CLEAR_ROW, column=5)
    if clear_cell.value and SOURCE_SHEET in str(clear_cell.value):
        clear_cell.value = None

    wb.save(dst_path)
    print(f"  Fixed: {src_path.name}")


def main():
    if not INPUT_DIR.exists():
        print(f"Error: Input folder '{INPUT_DIR}' not found.")
        return

    OUTPUT_DIR.mkdir(exist_ok=True)

    xlsx_files = list(INPUT_DIR.glob("*.xlsx")) + list(INPUT_DIR.glob("*.xlsm"))
    if not xlsx_files:
        print(f"No Excel files found in '{INPUT_DIR}'.")
        return

    print(f"Processing {len(xlsx_files)} file(s)...\n")
    for src in sorted(xlsx_files):
        dst = OUTPUT_DIR / src.name
        try:
            fix_file(src, dst)
        except Exception as e:
            print(f"  ERROR processing {src.name}: {e}")

    print(f"\nDone. Fixed files saved to '{OUTPUT_DIR}/'.")


if __name__ == "__main__":
    main()
